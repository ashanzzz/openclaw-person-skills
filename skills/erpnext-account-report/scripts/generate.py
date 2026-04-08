#!/root/.openclaw/workspace/.venv-xls/bin/python3
"""
ERPNext 账单导出 → 付款勾稽 → 明细拆分脚本

用法：
  python3 generate.py --company "天津祺富机械加工有限公司" \
      --account "现金账单-孟祥山 - 祺富" \
      --from "2026-01-01" --to "2026-01-31" \
      --output "/path/to/output.xlsx"

输出字段：
  凭证号 | 凭证日期 | 类型 | 供应商/对方 | 付款方式 | 项目名(物料名) |
  规格型号 | 单价 | 数量 | 单位 | 发票号码 | 分配金额 | 借贷方向 | 备注
"""

import argparse, requests, json, sys, re
from pathlib import Path

# ── API 基础 ────────────────────────────────────────────────────────────────
ENV_FILE = Path.home() / '.openclaw/workspace/secure/api-fillin.env'
SCRIPT_DIR = Path(__file__).parent

def load_env():
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env[k] = v
    return env

def api_get(path, params=None):
    env = load_env()
    token = env.get('ERP_API_TOKEN', '')
    base = env.get('ERP_BASE_URL', 'http://192.168.8.11:8888')
    headers = {'Authorization': f'Token {token}'}
    r = requests.get(f'{base}/api/resource/{path}', params=params, headers=headers, timeout=30)
    r.raise_for_status()
    return r.json().get('data', [])

def api_get_one(doctype, name):
    env = load_env()
    token = env.get('ERP_API_TOKEN', '')
    base = env.get('ERP_BASE_URL', 'http://192.168.8.11:8888')
    headers = {'Authorization': f'Token {token}'}
    r = requests.get(f'{base}/api/resource/{doctype}/{name}', headers=headers, timeout=30)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.json().get('data')

# ── GL Entry 查询 ────────────────────────────────────────────────────────────
def query_gl_entries(account, from_date, to_date):
    """返回指定账户和时间范围的 GL Entry 完整数据列表"""
    params = {
        'filters': json.dumps([
            ['account', '=', account],
            ['posting_date', 'between', [from_date, to_date]]
        ]),
        'order_by': 'posting_date asc',
        'limit_page_length': 10000
    }
    names = [x['name'] for x in api_get('GL Entry', params)]
    results = []
    for nm in names:
        d = api_get_one('GL Entry', nm)
        if d:
            results.append(d)
    return results

# ── Payment Entry 追查 ────────────────────────────────────────────────────────
def get_payment_entry_details(pe_name):
    """返回 PE 及其 references（含发票明细）"""
    pe = api_get_one('Payment Entry', pe_name)
    if not pe:
        return None, []
    # 收集该 PE 关联的所有 PI
    pi_names = set()
    for ref in pe.get('references', []):
        if ref.get('reference_doctype') == 'Purchase Invoice':
            pi_names.add(ref['reference_name'])
    pis = {}
    for pin in pi_names:
        pi = api_get_one('Purchase Invoice', pin)
        if pi:
            pis[pin] = pi
    return pe, pis

# ── Purchase Invoice 明细分拆 ────────────────────────────────────────────────
def get_invoice_items(pi):
    """返回 PI 下的 items 子表明细"""
    return pi.get('items', [])

# ── 金额分配 ────────────────────────────────────────────────────────────────
def allocate_by_items(pi_items, pi_total, allocated_amount):
    """按比例将分配金额分配到各物料行"""
    if pi_total == 0 or not pi_items:
        return [0.0] * len(pi_items)
    ratio = allocated_amount / pi_total
    return [round(item.get('amount', 0) * ratio, 2) for item in pi_items]

# ── 主处理 ─────────────────────────────────────────────────────────────────
def process_gl_entries(gl_entries, account_name):
    rows = []

    for gle in gl_entries:
        voucher_type = gle.get('voucher_type')
        voucher_no   = gle.get('voucher_no', '')
        posting_date = gle.get('posting_date', '')
        account      = gle.get('account', '')
        debit        = float(gle.get('debit', 0) or 0)
        credit       = float(gle.get('credit', 0) or 0)
        against      = gle.get('against', '')

        # 借贷方向
        direction = '支出' if credit > 0 else ('收入' if debit > 0 else '-')
        amount = credit if credit > 0 else debit

        if voucher_type == 'Payment Entry':
            pe, pis = get_payment_entry_details(voucher_no)
            if not pe:
                # 无关联PE，写一行汇总
                rows.append({
                    '凭证号': voucher_no,
                    '凭证日期': posting_date,
                    '类型': 'Payment Entry',
                    '供应商/对方': against,
                    '付款方式': pe.get('mode_of_payment') if pe else '-',
                    '项目名(物料名)': '(无法获取明细)',
                    '规格型号': '-',
                    '单价': 0,
                    '数量': 0,
                    '单位': '-',
                    '发票号码': '-',
                    '分配金额': amount,
                    '借贷方向': direction,
                    '备注': gle.get('remarks', '')[:100]
                })
                continue

            mode_of_payment = pe.get('mode_of_payment', '-')

            # 按 references 遍历
            for ref in pe.get('references', []):
                if ref.get('reference_doctype') != 'Purchase Invoice':
                    continue
                pin = ref.get('reference_name')
                allocated = float(ref.get('allocated_amount', 0) or 0)
                bill_no   = ref.get('bill_no', '')

                pi = pis.get(pin)
                if not pi:
                    rows.append({
                        '凭证号': voucher_no,
                        '凭证日期': posting_date,
                        '类型': 'Payment Entry',
                        '供应商/对方': against,
                        '付款方式': mode_of_payment,
                        '项目名(物料名)': f'(PI不存在: {pin})',
                        '规格型号': '-',
                        '单价': 0, '数量': 0, '单位': '-',
                        '发票号码': bill_no,
                        '分配金额': allocated,
                        '借贷方向': direction,
                        '备注': gle.get('remarks', '')[:100]
                    })
                    continue

                # 获取 PI 明细
                pi_items  = get_invoice_items(pi)
                pi_total  = float(pi.get('grand_total', 0) or 0)
                allocated_list = allocate_by_items(pi_items, pi_total, allocated)

                for item, alloc in zip(pi_items, allocated_list):
                    rows.append({
                        '凭证号': voucher_no,
                        '凭证日期': posting_date,
                        '类型': 'Payment Entry',
                        '供应商/对方': against,
                        '付款方式': mode_of_payment,
                        '项目名(物料名)': item.get('item_name', '-'),
                        '规格型号': item.get('custom_guige_xinghao', '-'),
                        '单价': round(float(item.get('rate', 0) or 0), 4),
                        '数量': round(float(item.get('qty', 0) or 0), 3),
                        '单位': item.get('uom', '-'),
                        '发票号码': bill_no,
                        '分配金额': alloc,
                        '借贷方向': direction,
                        '备注': item.get('custom_备注', '')
                    })

        elif voucher_type == 'Purchase Invoice':
            # 直接 PI（无PE），如房租/电费
            pi = api_get_one('Purchase Invoice', voucher_no)
            if not pi:
                rows.append({
                    '凭证号': voucher_no,
                    '凭证日期': posting_date,
                    '类型': 'Purchase Invoice',
                    '供应商/对方': against,
                    '付款方式': '-',
                    '项目名(物料名)': f'(PI不存在: {voucher_no})',
                    '规格型号': '-',
                    '单价': 0, '数量': 0, '单位': '-',
                    '发票号码': '-',
                    '分配金额': amount,
                    '借贷方向': direction,
                    '备注': gle.get('remarks', '')[:100]
                })
                continue

            pi_items = get_invoice_items(pi)
            pi_total = float(pi.get('grand_total', 0) or 0)
            bill_no  = pi.get('bill_no', '')

            if not pi_items:
                # 无items，写一行汇总
                rows.append({
                    '凭证号': voucher_no,
                    '凭证日期': posting_date,
                    '类型': 'Purchase Invoice',
                    '供应商/对方': against,
                    '付款方式': '-',
                    '项目名(物料名)': pi.get('supplier', '-'),
                    '规格型号': '-',
                    '单价': amount, '数量': 1, '单位': '-',
                    '发票号码': bill_no,
                    '分配金额': amount,
                    '借贷方向': direction,
                    '备注': pi.get('custom_备注', '')
                })
            else:
                allocated_list = allocate_by_items(pi_items, pi_total, amount)
                for item, alloc in zip(pi_items, allocated_list):
                    rows.append({
                        '凭证号': voucher_no,
                        '凭证日期': posting_date,
                        '类型': 'Purchase Invoice',
                        '供应商/对方': against,
                        '付款方式': '-',
                        '项目名(物料名)': item.get('item_name', '-'),
                        '规格型号': item.get('custom_guige_xinghao', '-'),
                        '单价': round(float(item.get('rate', 0) or 0), 4),
                        '数量': round(float(item.get('qty', 0) or 0), 3),
                        '单位': item.get('uom', '-'),
                        '发票号码': bill_no,
                        '分配金额': alloc,
                        '借贷方向': direction,
                        '备注': item.get('custom_备注', '')
                    })
        else:
            # 其他类型（Journal Entry等），写汇总行
            rows.append({
                '凭证号': voucher_no,
                '凭证日期': posting_date,
                '类型': voucher_type,
                '供应商/对方': against,
                '付款方式': '-',
                '项目名(物料名)': '(非采购类型)',
                '规格型号': '-',
                '单价': 0, '数量': 0, '单位': '-',
                '发票号码': '-',
                '分配金额': amount,
                '借贷方向': direction,
                '备注': gle.get('remarks', '')[:100]
            })

    return rows

# ── Excel 输出 ─────────────────────────────────────────────────────────────
def to_excel(rows, output_path, account_name, from_date, to_date):
    try:
        import openpyxl
    except ImportError:
        sys.path.insert(0, str(SCRIPT_DIR.parent / '.venv-xls/lib/python3.12/site-packages'))
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '账单明细'

    hf   = PatternFill('solid', fgColor='1F4E79')
    hfont= Font(bold=True, color='FFFFFF', size=11)
    tfont= Font(bold=True, size=13)
    ctr  = Alignment(horizontal='center', vertical='center')
    rgt  = Alignment(horizontal='right', vertical='center')
    thin = Side(style='thin')
    bd   = Border(top=thin, bottom=thin, left=thin, right=thin)
    alt  = PatternFill('solid', fgColor='EBF3FB')

    cols = ['凭证号','凭证日期','类型','供应商/对方','付款方式',
            '项目名(物料名)','规格型号','单价','数量','单位',
            '发票号码','分配金额','借贷方向','备注']

    ws['A1'] = f'账单明细：{account_name}'
    ws['A1'].font = tfont; ws['A1'].alignment = ctr
    ws.merge_cells('A1:N1')
    ws.row_dimensions[1].height = 22

    ws['A2'] = f'{from_date} ~ {to_date}  |  共 {len(rows)} 条'
    ws['A2'].alignment = ctr; ws.merge_cells('A2:N2')

    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = ctr; cell.border = bd
    ws.row_dimensions[3].height = 18

    for i, row in enumerate(rows, 1):
        fill = alt if i % 2 == 0 else None
        for c, col in enumerate(cols, 1):
            val = row.get(col, '')
            cell = ws.cell(row=i+3, column=c, value=val)
            cell.border = bd
            if fill: cell.fill = fill
            if col in ('凭证号','凭证日期','类型','供应商/对方','付款方式','单位','发票号码','借贷方向','备注'):
                cell.alignment = ctr
            if col in ('单价','分配金额'):
                cell.alignment = rgt
                cell.number_format = '#,##0.0000' if col == '单价' else '#,##0.00'
            if col == '数量':
                cell.alignment = rgt; cell.number_format = '#,##0.000'

    # 合计行
    lr = len(rows) + 4
    total_debit  = sum(float(r.get('分配金额', 0)) for r in rows if r.get('借贷方向') == '收入')
    total_credit = sum(float(r.get('分配金额', 0)) for r in rows if r.get('借贷方向') == '支出')
    ws.cell(row=lr, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=lr, column=12, value=total_credit).number_format = '#,##0.00'
    ws.cell(row=lr, column=12).font = Font(bold=True); ws.cell(row=lr, column=12).alignment = rgt
    for c in range(1, len(cols)+1):
        ws.cell(row=lr, column=c).border = bd
        ws.cell(row=lr, column=c).fill = PatternFill('solid', fgColor='D9E1F2')

    widths = [22, 12, 16, 18, 10, 22, 14, 12, 10, 6, 20, 12, 8, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A4'

    wb.save(output_path)
    print(f'✅ 已保存：{output_path}')
    print(f'   共 {len(rows)} 条 | 支出 ¥{total_credit:,.2f} | 收入 ¥{total_debit:,.2f}')

# ── 入口 ──────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='ERPNext 账单导出 → 付款勾稽 → 明细分拆')
    parser.add_argument('--company',  required=True, help='公司名称')
    parser.add_argument('--account',  required=True, help='账户名称（完整名称，含 - 公司后缀）')
    parser.add_argument('--from',     required=True, dest='from_date', help='起始日期 YYYY-MM-DD')
    parser.add_argument('--to',       required=True, dest='to_date',   help='截止日期 YYYY-MM-DD')
    parser.add_argument('--output',  required=True, help='输出 xlsx 路径')
    args = parser.parse_args()

    print(f'查询账户：{args.account}')
    print(f'日期范围：{args.from_date} ~ {args.to_date}')
    gl_entries = query_gl_entries(args.account, args.from_date, args.to_date)
    print(f'GL 条数：{len(gl_entries)}')

    rows = process_gl_entries(gl_entries, args.account)
    to_excel(rows, args.output, args.account, args.from_date, args.to_date)

if __name__ == '__main__':
    main()

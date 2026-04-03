#!/usr/bin/env python3
"""Generate invoice Excel - uses curl for HTTP

注意：发票项目名称的「总结类/明细类」判断由 Agent 智能执行，
不要在此脚本中硬编码关键词替换规则。
"""
import subprocess, json, csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

ERP_KEY = "d8837cf53a3cfae"
ERP_SECRET = "2a26d85166a23f3"
ERP_URL = "http://192.168.8.11:8888"

def curl_get(resource_path):
    url = ERP_URL + "/api/resource/" + resource_path
    cmd = ["curl", "-s", url, "-H", "Authorization: Token " + ERP_KEY + ":" + ERP_SECRET]
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
    try:
        return json.loads(r.stdout)
    except:
        return {}

def inv_type(rate):
    if rate is None or rate == 0: return "定额发票"
    elif abs(rate - 13.0) < 0.01: return "专用发票"
    else: return "普通发票"

def pur_type(exp):
    if not exp: return "费用"
    if "暂估" in exp or "库存" in exp: return "购买申请"
    elif "管理" in exp: return "报销申请"
    return "费用"

def date_serial(d):
    if not d: return None
    try:
        return (datetime.strptime(d[:10], "%Y-%m-%d") - datetime(1899, 12, 30)).days
    except:
        return None

def make_excel(rows, month_str, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "发票明细"
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value=month_str + "发票明细")
    c.font = Font(name="等线", size=12, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    ws.row_dimensions[1].height = 28
    headers = ["序号","发票号码","发票日期","发票类型","发票项目","金额","购买方式","支付方式"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="等线", size=10); c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    for ci, w in {1:4.5,2:16,3:13,4:7,5:18,6:11,7:8,8:7}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, row in enumerate(rows, 3):
        vals = [
            (1, row["seq"], "center", None),
            (2, row["bill_no"], "center", "@"),
            (3, row["dt"], "center", 'yyyy"年"mm"月"dd"日"'),
            (4, row["inv_t"], "center", None),
            (5, row["nm"], "center", None),
            (6, row["gr"], "center", "0.00"),
            (7, row["pt"], "center", None),
            (8, "现金", "center", None),
        ]
        for ci, val, aln, fmt in vals:
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="等线", size=9); c.border = border; c.alignment = Alignment(horizontal=aln)
            if fmt: c.number_format = fmt
        ws.row_dimensions[ri].height = 16.5
    wb.save(out_path)
    print(f"Saved {out_path}: {len(rows)} rows")

def build_rows(csv_path):
    pis = {}
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if int(row["docstatus"]) == 1:
                pis[row["name"]] = row

    rows = []
    total = len(pis)
    fields_enc = "%5B%22items%22%2C%22taxes%22%2C%22custom_%E5%8F%91%E7%A5%A8%E7%B1%BB%E5%9E%8B%22%5D"
    for i, (name, pi) in enumerate(pis.items()):
        print(f"  [{i+1}/{total}] {name}...")
        d = curl_get(f"Purchase%20Invoice/{name}?fields={fields_enc}")
        if "data" not in d:
            print(f"    失败")
            continue
        items = d["data"].get("items", [])
        taxes = d["data"].get("taxes", [])
        custom_type = d["data"].get("custom_发票类型")
        posting_date = pi.get("posting_date", "")
        if not items: continue
        tr = float(taxes[0].get("rate", 0)) if taxes else 0.0
        for it in items:
            exp = it.get("expense_account", "")
            gross = float(it.get("amount", 0))
            raw_nm = it.get("item_name") or it.get("description") or ""
            # item_name 直接写入，由 Agent 提前做总结类/明细类判断
            rows.append({
                "seq": len(rows) + 1,
                "bill_no": pi.get("bill_no") or "",
                "dt": date_serial(posting_date),
                "inv_t": custom_type if custom_type else inv_type(tr),
                "nm": raw_nm,
                "gr": round(gross, 2),
                "pt": pur_type(exp),
            })
    return rows

if __name__ == "__main__":
    # 祺富
    print("=== 祺富 ===")
    qifu_rows = build_rows(
        "/root/.openclaw/workspace/secure/erpnext/exports/天津祺富机械加工有限公司/2026-03/invoices_Purchase_Invoice_2026-03.csv")
    print(f"祺富: {len(qifu_rows)} 行")
    make_excel(qifu_rows, "2026年03月", "/root/.openclaw/workspace/祺富_2026-03_发票明细.xlsx")

    # 吉众
    print("\n=== 吉众 ===")
    jizhong_rows = build_rows(
        "/root/.openclaw/workspace/secure/erpnext/exports/天津吉众科技有限公司/2026-03/invoices_Purchase_Invoice_2026-03.csv")
    print(f"吉众: {len(jizhong_rows)} 行")
    make_excel(jizhong_rows, "2026年03月", "/root/.openclaw/workspace/吉众_2026-03_发票明细.xlsx")

    print("\n全部完成")

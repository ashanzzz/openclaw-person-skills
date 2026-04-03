#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Analyze recurring no-invoice cash expenses from a marked ledger workbook.

Default input is 2025-12 marked ledger (用户提供：历史带标记账单...).
Focus:
- Sheet: 没有发票现金账
- Year: 2025

Outputs (to workspace outbox):
- no_invoice_cash_2025_stats.csv
- no_invoice_cash_2025_summary.md

Safe: read-only.
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict

import openpyxl

WORKSPACE = Path("/root/.openclaw/workspace")
DEFAULT_IN = WORKSPACE / "media/inbound/marked_ledger_baseline.xlsx"  # optional symlink/copy
OUT_DIR = WORKSPACE / "outbox"


def to_date(v: Any) -> datetime | None:
    if isinstance(v, datetime):
        return v
    return None


def norm_item(s: str) -> str:
    t = (s or "").strip()
    t = t.replace("（", "(").replace("）", ")")
    t = re.sub(r"^[\*＊]+", "", t).strip()
    t = re.sub(r"^(\d{1,2})月", "", t).strip()
    t = re.sub(r"^(20\d{2})年(\d{1,2})月", "", t).strip()
    t = re.sub(r"\s+", " ", t).strip()
    t = t.replace("现金账帐转入", "现金账转入")
    return t


def norm_remark(s: str) -> str:
    t = (s or "").strip()
    t = t.replace("（", "(").replace("）", ")")
    t = re.sub(r"\s+", " ", t).strip()
    return t


@dataclass
class Agg:
    count: int = 0
    months: set[str] = None  # type: ignore
    total_spend: float = 0.0
    total_income: float = 0.0
    remarks: Dict[str, int] = None  # type: ignore

    def __post_init__(self):
        self.months = set()
        self.remarks = defaultdict(int)


def analyze(xlsx_path: Path, year: int = 2025, sheet: str = "没有发票现金账"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet not found: {sheet}")

    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx for idx, name in enumerate(header) if name}

    def get(row, key):
        i = col_map.get(key)
        return None if i is None else row[i]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = to_date(get(r, "日期"))
        if not d or d.year != year:
            continue
        item = str(get(r, "项目") or "").strip()
        income = float(get(r, "收入") or 0 or 0)
        spend = float(get(r, "支出") or 0 or 0)
        remark = str(get(r, "备注") or "").strip()
        rows.append((d, item, income, spend, remark))

    aggs: Dict[str, Agg] = {}
    for d, item, income, spend, remark in rows:
        k = norm_item(item)
        if not k:
            continue
        a = aggs.setdefault(k, Agg())
        a.count += 1
        a.months.add(d.strftime("%Y-%m"))
        a.total_income += income
        a.total_spend += spend
        rr = norm_remark(remark)
        if rr:
            a.remarks[rr] += 1

    items = sorted(
        aggs.items(),
        key=lambda kv: (len(kv[1].months), kv[1].total_spend, kv[1].count),
        reverse=True,
    )
    return rows, items


def main():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(DEFAULT_IN), help="marked ledger xlsx path")
    ap.add_argument("--year", type=int, default=2025)
    ap.add_argument("--sheet", default="没有发票现金账")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"missing xlsx: {xlsx_path}")

    rows, items = analyze(xlsx_path, year=args.year, sheet=args.sheet)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    csv_path = OUT_DIR / f"no_invoice_cash_{args.year}_stats.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["项目(归一)", "出现次数", "覆盖月份数", "月份列表", "支出合计", "收入合计", "常见备注Top3"])
        for k, a in items:
            top_remarks = sorted(a.remarks.items(), key=lambda x: x[1], reverse=True)[:3]
            top_remarks_s = " | ".join([f"{t}({n})" for t, n in top_remarks])
            months = ",".join(sorted(a.months))
            w.writerow([k, a.count, len(a.months), months, round(a.total_spend, 2), round(a.total_income, 2), top_remarks_s])

    md_path = OUT_DIR / f"no_invoice_cash_{args.year}_summary.md"
    recurring = [(k, a) for k, a in items if len(a.months) >= 6 and a.total_spend > 0]
    top_spend = sorted(items, key=lambda kv: kv[1].total_spend, reverse=True)[:20]

    lines = []
    lines.append(f"# {args.year}年 无发票现金账 规律统计\n")
    lines.append(f"- 来源文件：`{xlsx_path}`\n")
    lines.append(f"- 统计口径：sheet=`{args.sheet}`\n")
    lines.append(f"- 行数：{len(rows)}\n")
    lines.append(f"- 归一项目数：{len(items)}\n")

    lines.append("\n## A. 每月高频（覆盖≥6个月，且支出>0）\n")
    for k, a in recurring[:30]:
        top_remarks = sorted(a.remarks.items(), key=lambda x: x[1], reverse=True)[:2]
        lines.append(
            f"- {k}｜覆盖{len(a.months)}个月｜支出{a.total_spend:.2f}｜次数{a.count}｜备注例：" + "; ".join([t for t, _ in top_remarks])
        )

    lines.append("\n## B. 支出金额Top20\n")
    for k, a in top_spend:
        lines.append(f"- {k}｜支出{a.total_spend:.2f}｜覆盖{len(a.months)}个月｜次数{a.count}")

    md_path.write_text("\n".join(lines), encoding="utf-8")

    print("OK")
    print("csv", csv_path)
    print("md", md_path)


if __name__ == "__main__":
    main()
